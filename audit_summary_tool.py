"""
Audit Summary Tool
==================

Reads multiple site-audit Excel workbooks from a folder (e.g. files named like
"Audit_SUM-SU-STB-1326_20260521.xlsx"), pulls the DRM / After-MOCN antenna
values (Antenna Type, Antenna Height, Azimuth, M-tilt) per sector from the
"L2100 + 1800+ L850" sheet, compares DRM vs After MOCN, and appends the
summary as new rows at the END of an existing tracker workbook
(tracker_audit.xlsx) -- existing rows are never touched or overwritten.

If a site's "L2300" sheet also has complete (non-N/A) values, extra rows are
appended right after that site's main rows, tagged with REMARK = "L2300".

Usage:
    python audit_summary_tool.py

This opens a small GUI:
  1. Browse Folder  -> folder containing the source audit .xlsx files
  2. Browse Tracker  -> the existing tracker_audit.xlsx to append into
  3. Run             -> processes every workbook in the folder and appends
                         results to the tracker file (a timestamped backup
                         of the tracker is made first).

All the cell/row mapping below matches the layout observed in the sample
files. If a future source file uses a different layout, adjust the CONFIG
section only -- the rest of the logic does not need to change.
"""

import os
import sys
import glob
import shutil
import queue
import threading
import traceback
from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

# ----------------------------------------------------------------------------
# CONFIG -- cell/row mapping (edit here if a source file layout changes)
# ----------------------------------------------------------------------------

# Sheet names in the SOURCE audit workbook (matched with .strip() so trailing
# spaces like "L2100 + 1800+ L850 " don't matter).
SRC_SHEET_MAIN = "L2100 + 1800+ L850"
SRC_SHEET_L2300 = "L2300"

# New Site ID lives at B3 (merged B3:C3) on the main sheet.
SITE_ID_CELL = "B3"

# Sector number row, one value per sector column (D5:E5, F5:G5, H5:I5, J5:K5
# are all merged pairs -- we only need the left column of each pair).
SECTOR_ROW = 5
# up to 4 sectors; invalid/N/A ones are skipped
SECTOR_COLS = ["D", "F", "H", "J"]

# Row numbers for each metric on the main sheet.
AZIMUTH_DRM_ROW = 6          # "Orientation(*) Planning"  -> DRM
AZIMUTH_AFTER_ROW = 7        # "Orientation(*) After"     -> After MOCN
MTILT_DRM_ROW = 8            # "Mech Tilt(*) Planning"    -> DRM
MTILT_AFTER_ROW = 9          # "Mech Tilt(*) After"       -> After MOCN
ANTENNA_TYPE_ROW = 10        # "Model" (shared by DRM & After MOCN)
# "Antenna Height (M)" (shared by DRM & After MOCN)
ANTENNA_HEIGHT_ROW = 11

# Tracker (output) workbook layout.
TRACKER_SHEET_NAME = "Tracker Audit"
TRACKER_FIRST_DATA_ROW = 3   # rows 1-2 are headers
TRACKER_COLS = {
    "site_id": "A",
    "sector": "B",
    "drm_type": "C",
    "drm_height": "D",
    "drm_azimuth": "E",
    "drm_mtilt": "F",
    "after_type": "G",
    "after_height": "H",
    "after_azimuth": "I",
    "after_mtilt": "J",
    "status": "K",
    "remark": "L",
}

NA_VALUES = {"", "N/A", "NA", "-", "NONE"}


# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------

def is_valid(value):
    """True if a cell value counts as real data (not blank / N/A)."""
    if value is None:
        return False
    return str(value).strip().upper() not in NA_VALUES


def read_merged(ws, coord):
    """
    Read a cell's value, resolving merged cells: if `coord` falls inside a
    merged range, return the value stored at the top-left anchor cell of
    that range (openpyxl only stores the value there).
    """
    cell = ws[coord]
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            min_col, min_row, _, _ = merged_range.bounds
            return ws.cell(row=min_row, column=min_col).value
    return cell.value


def find_sheet(wb, target_name):
    """Find a sheet whose name matches target_name, ignoring surrounding whitespace."""
    target = target_name.strip().lower()
    for name in wb.sheetnames:
        if name.strip().lower() == target:
            return wb[name]
    return None


def get_row_values(ws, row, cols):
    """Return {col: value} for the given row/cols, resolving merged cells."""
    return {c: read_merged(ws, f"{c}{row}") for c in cols}


def build_status_remark(drm_azimuth, after_azimuth, drm_mtilt, after_mtilt,
                        drm_type, after_type, drm_height, after_height,
                        extra_tag=None):
    """
    Compare DRM vs After MOCN values and produce (STATUS, REMARK).
    STATUS is MATCH if all four fields are unchanged, else MISMATCH.
    REMARK lists which field(s) changed; extra_tag (e.g. "L2300") is
    prepended when set.
    """
    diffs = []
    pairs = [
        ("Antenna Type", drm_type, after_type),
        ("Antenna Height", drm_height, after_height),
        ("Azimuth", drm_azimuth, after_azimuth),
        ("M-tilt", drm_mtilt, after_mtilt),
    ]
    for label, before, after in pairs:
        if str(before).strip() != str(after).strip():
            diffs.append(f"{label} changed: {before} -> {after}")

    status = "MISMATCH" if diffs else "MATCH"
    remark_parts = []
    if extra_tag:
        remark_parts.append(extra_tag)
    remark_parts.extend(diffs)
    remark = " | ".join(remark_parts) if remark_parts else ""
    return status, remark


# ----------------------------------------------------------------------------
# Core extraction
# ----------------------------------------------------------------------------

def extract_sheet_rows(ws, site_id, extra_tag=None):
    """
    Extract one row per valid sector from a sheet that follows the
    DRM/After-MOCN layout (used for both the main sheet and L2300).
    Returns a list of dicts matching TRACKER_COLS keys (minus site_id/sector
    filled in here too).
    """
    sector_vals = get_row_values(ws, SECTOR_ROW, SECTOR_COLS)
    azimuth_drm = get_row_values(ws, AZIMUTH_DRM_ROW, SECTOR_COLS)
    azimuth_after = get_row_values(ws, AZIMUTH_AFTER_ROW, SECTOR_COLS)
    mtilt_drm = get_row_values(ws, MTILT_DRM_ROW, SECTOR_COLS)
    mtilt_after = get_row_values(ws, MTILT_AFTER_ROW, SECTOR_COLS)
    ant_type = get_row_values(ws, ANTENNA_TYPE_ROW, SECTOR_COLS)
    ant_height = get_row_values(ws, ANTENNA_HEIGHT_ROW, SECTOR_COLS)

    rows = []
    for col in SECTOR_COLS:
        # A sector counts as real if it has a valid azimuth (planning) or
        # a valid antenna type -- both being N/A means the sector doesn't
        # exist on this site/band.
        if not (is_valid(azimuth_drm[col]) or is_valid(ant_type[col])):
            continue

        status, remark = build_status_remark(
            azimuth_drm[col], azimuth_after[col],
            mtilt_drm[col], mtilt_after[col],
            ant_type[col], ant_type[col],       # antenna type shared
            ant_height[col], ant_height[col],   # antenna height shared
            extra_tag=extra_tag,
        )

        rows.append({
            "site_id": site_id,
            "sector": sector_vals[col],
            "drm_type": ant_type[col],
            "drm_height": ant_height[col],
            "drm_azimuth": azimuth_drm[col],
            "drm_mtilt": mtilt_drm[col],
            "after_type": ant_type[col],
            "after_height": ant_height[col],
            "after_azimuth": azimuth_after[col],
            "after_mtilt": mtilt_after[col],
            "status": status,
            "remark": remark,
        })
    return rows


def process_file(path, log=print):
    """
    Process a single source workbook. Returns a list of row dicts
    (main-sheet rows followed by any L2300 rows), or [] if the file
    doesn't have the expected main sheet.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    ws_main = find_sheet(wb, SRC_SHEET_MAIN)
    if ws_main is None:
        log(f"  [skip] '{os.path.basename(path)}': no sheet named "
            f"'{SRC_SHEET_MAIN}' found (sheets: {wb.sheetnames})")
        return []

    site_id = read_merged(ws_main, SITE_ID_CELL)
    if not is_valid(site_id):
        log(f"  [skip] '{os.path.basename(path)}': Site ID cell "
            f"{SITE_ID_CELL} on '{ws_main.title}' is empty")
        return []

    rows = extract_sheet_rows(ws_main, site_id)
    log(f"  '{os.path.basename(path)}' -> site {site_id}: "
        f"{len(rows)} sector row(s) from '{ws_main.title}'")

    ws_l2300 = find_sheet(wb, SRC_SHEET_L2300)
    if ws_l2300 is not None:
        l2300_rows = extract_sheet_rows(ws_l2300, site_id, extra_tag="L2300")
        if l2300_rows:
            log(f"    + L2300 sheet has complete value(s): "
                f"adding {len(l2300_rows)} extra row(s), tagged L2300")
            rows.extend(l2300_rows)

    return rows


def process_folder(folder, tracker_path, log=print):
    """Process every .xlsx in `folder` (skipping temp/lock files and the
    tracker file itself) and append the results to the tracker workbook."""
    pattern = os.path.join(folder, "*.xlsx")
    files = sorted(glob.glob(pattern))

    tracker_abs = os.path.abspath(tracker_path)
    files = [
        f for f in files
        if not os.path.basename(f).startswith("~$")
        and os.path.abspath(f) != tracker_abs
    ]

    if not files:
        log(f"No .xlsx files found in: {folder}")
        return 0

    log(f"Found {len(files)} file(s) in {folder}")
    all_rows = []
    for path in files:
        try:
            rows = process_file(path, log=log)
            all_rows.extend(rows)
        except Exception as exc:
            log(f"  [error] '{os.path.basename(path)}': {exc}")
            log(traceback.format_exc())

    if not all_rows:
        log("No rows extracted from any file -- tracker not modified.")
        return 0

    append_to_tracker(tracker_path, all_rows, log=log)
    return len(all_rows)


# ----------------------------------------------------------------------------
# Tracker append (never touches existing rows)
# ----------------------------------------------------------------------------

def find_next_empty_row(ws, first_data_row):
    """Scan column A from the bottom to find the first row after the
    existing data (robust against ws.max_row over-counting formatted-but-
    empty rows)."""
    row = ws.max_row
    while row >= first_data_row and ws.cell(row=row, column=1).value in (None, ""):
        row -= 1
    return max(row + 1, first_data_row)


def append_to_tracker(tracker_path, rows, log=print):
    # Safety backup before writing.
    backup_path = (
        f"{os.path.splitext(tracker_path)[0]}_backup_"
        f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
    shutil.copy2(tracker_path, backup_path)
    log(f"Backup of tracker saved: {os.path.basename(backup_path)}")

    wb = openpyxl.load_workbook(tracker_path)
    ws = wb[TRACKER_SHEET_NAME] if TRACKER_SHEET_NAME in wb.sheetnames else wb.active

    next_row = find_next_empty_row(ws, TRACKER_FIRST_DATA_ROW)
    log(f"Appending {len(rows)} row(s) starting at row {next_row} "
        f"(existing rows above are left untouched)")

    for i, row in enumerate(rows):
        r = next_row + i
        for key, col_letter in TRACKER_COLS.items():
            col_idx = column_index_from_string(col_letter)
            ws.cell(row=r, column=col_idx, value=row.get(key))

    wb.save(tracker_path)
    log(f"Saved: {tracker_path}")


# ----------------------------------------------------------------------------
# GUI
# ----------------------------------------------------------------------------

def launch_gui():
    import tkinter as tk
    from tkinter import filedialog, scrolledtext, messagebox

    root = tk.Tk()
    root.title("AOR Summary")
    root.geometry("720x480")

    folder_var = tk.StringVar()
    tracker_var = tk.StringVar()
    log_queue = queue.Queue()

    def browse_folder():
        path = filedialog.askdirectory(
            title="Select folder with audit .xlsx files")
        if path:
            folder_var.set(path)

    def browse_tracker():
        path = filedialog.askopenfilename(
            title="Select tracker .xlsx",
            filetypes=[("Excel files", "*.xlsx")],
        )
        if path:
            tracker_var.set(path)

    def log(msg):
        log_queue.put(str(msg))

    def poll_log():
        try:
            while True:
                msg = log_queue.get_nowait()
                log_box.configure(state="normal")
                log_box.insert(tk.END, msg + "\n")
                log_box.see(tk.END)
                log_box.configure(state="disabled")
        except queue.Empty:
            pass
        root.after(100, poll_log)

    def run_worker():
        folder = folder_var.get().strip()
        tracker = tracker_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showerror(
                "Error", "Please choose a valid source folder.")
            return
        if not tracker or not os.path.isfile(tracker):
            messagebox.showerror(
                "Error", "Please choose a valid tracker .xlsx file.")
            return

        run_btn.configure(state="disabled")

        def task():
            log(f"--- Run started {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} ---")
            try:
                count = process_folder(folder, tracker, log=log)
                log(f"--- Done: {count} row(s) appended ---")
            except Exception as exc:
                log(f"[FATAL] {exc}")
                log(traceback.format_exc())
            finally:
                run_btn.configure(state="normal")

        threading.Thread(target=task, daemon=True).start()

    # Layout
    pad = {"padx": 8, "pady": 6}

    frame_top = tk.Frame(root)
    frame_top.pack(fill="x", **pad)

    tk.Label(frame_top, text="Source Folder (audit files):",
             width=24, anchor="w").grid(row=0, column=0, sticky="w")
    tk.Entry(frame_top, textvariable=folder_var).grid(
        row=0, column=1, sticky="ew", padx=4)
    tk.Button(frame_top, text="Browse Folder",
              command=browse_folder).grid(row=0, column=2)

    tk.Label(frame_top, text="Tracker File (output .xlsx):",
             width=24, anchor="w").grid(row=1, column=0, sticky="w")
    tk.Entry(frame_top, textvariable=tracker_var).grid(
        row=1, column=1, sticky="ew", padx=4)
    tk.Button(frame_top, text="Browse Tracker",
              command=browse_tracker).grid(row=1, column=2)

    frame_top.columnconfigure(1, weight=1)

    run_btn = tk.Button(
        root, text="Run", command=run_worker, width=20, height=2)
    run_btn.pack(pady=6)

    log_box = scrolledtext.ScrolledText(root, state="disabled", wrap="word")
    log_box.pack(fill="both", expand=True, padx=8, pady=8)

    root.after(100, poll_log)
    root.mainloop()


if __name__ == "__main__":
    launch_gui()
